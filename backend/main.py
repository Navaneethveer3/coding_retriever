from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Header, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone
import openpyxl
import io
import os

from database import get_db, init_db, engine, Base
from models import User, Student
from auth import (
    verify_password,
    get_password_hash,
    create_access_token,
    decode_access_token,
)
from scraper import scrape_student_data

app = FastAPI(title="Coding Retriever", version="1.0.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Pydantic Schemas ───────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str


class CreateAccountRequest(BaseModel):
    username: str
    password: str


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class UpdateUserRequest(BaseModel):
    username: str
    password: Optional[str] = None


class UserResponse(BaseModel):
    id: int
    username: str
    is_admin: bool
    has_uploaded_1st_year: bool
    has_uploaded_2nd_year: bool
    has_uploaded_3rd_year: bool
    has_uploaded_4th_year: bool

    class Config:
        from_attributes = True


# ─── Helpers ─────────────────────────────────────────────────────
def get_current_user(authorization: str = Header(None), db: Session = Depends(get_db)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = authorization.split(" ")[1]
    payload = decode_access_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    user = db.query(User).filter(User.username == payload.get("sub")).first()
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


# ─── Startup ─────────────────────────────────────────────────────
@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)
    db = next(get_db())
    # Create default admin if not exists
    admin = db.query(User).filter(User.username == "Admin@AI").first()
    if not admin:
        admin = User(
            username="Admin@AI",
            password_hash=get_password_hash("AI@Artificial_Intelligence"),
            is_admin=True,
        )
        db.add(admin)
        db.commit()
    db.close()


# ─── Auth Routes ─────────────────────────────────────────────────
@app.post("/api/auth/login")
def login(req: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == req.username).first()
    if not user or not verify_password(req.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid username or password")

    token = create_access_token(data={"sub": user.username, "is_admin": user.is_admin})
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "id": user.id,
            "username": user.username,
            "is_admin": user.is_admin,
            "has_uploaded_1st_year": user.has_uploaded_1st_year,
            "has_uploaded_2nd_year": user.has_uploaded_2nd_year,
            "has_uploaded_3rd_year": user.has_uploaded_3rd_year,
            "has_uploaded_4th_year": user.has_uploaded_4th_year,
        },
    }


@app.post("/api/auth/create-account")
def create_account(
    req: CreateAccountRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Only admin can create accounts")

    existing = db.query(User).filter(User.username == req.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")

    new_user = User(
        username=req.username,
        password_hash=get_password_hash(req.password),
        is_admin=False,
    )
    db.add(new_user)
    db.commit()
    return {"message": "Account created successfully", "username": req.username}


@app.get("/api/admin/users", response_model=List[UserResponse])
def list_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    return db.query(User).all()


@app.put("/api/admin/users/{user_id}")
def update_user(
    user_id: int,
    req: UpdateUserRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    user_to_update = db.query(User).filter(User.id == user_id).first()
    if not user_to_update:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Check if new username is taken by someone else
    if req.username != user_to_update.username:
        existing = db.query(User).filter(User.username == req.username).first()
        if existing:
            raise HTTPException(status_code=400, detail="Username already taken")
        user_to_update.username = req.username
    
    if req.password:
        user_to_update.password_hash = get_password_hash(req.password)
        
    db.commit()
    return {"message": "User updated successfully"}


@app.delete("/api/admin/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
        
    user_to_delete = db.query(User).filter(User.id == user_id).first()
    if not user_to_delete:
        raise HTTPException(status_code=404, detail="User not found")
        
    db.delete(user_to_delete)
    db.commit()
    return {"message": "User deleted successfully"}


# ─── Profile Routes ─────────────────────────────────────────────
@app.get("/api/profile")
def get_profile(current_user: User = Depends(get_current_user)):
    return {
        "id": current_user.id,
        "username": current_user.username,
        "is_admin": current_user.is_admin,
        "has_uploaded_1st_year": current_user.has_uploaded_1st_year,
        "has_uploaded_2nd_year": current_user.has_uploaded_2nd_year,
        "has_uploaded_3rd_year": current_user.has_uploaded_3rd_year,
        "has_uploaded_4th_year": current_user.has_uploaded_4th_year,
    }


@app.post("/api/profile/change-password")
def change_password(
    req: ChangePasswordRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not verify_password(req.current_password, current_user.password_hash):
        raise HTTPException(status_code=400, detail="Current password is incorrect")

    current_user.password_hash = get_password_hash(req.new_password)
    db.commit()
    return {"message": "Password changed successfully"}


# ─── Upload Routes ───────────────────────────────────────────────
@app.post("/api/upload/{category}")
async def upload_excel(
    category: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if category not in ("1st_year", "2nd_year", "3rd_year", "4th_year"):
        raise HTTPException(status_code=400, detail="Category must be 1st, 2nd, 3rd, or 4th year")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed")

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        sheet = workbook.active

        # Find column indices from header row
        headers = {}
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value:
                header_lower = str(cell.value).strip().lower()
                if "name" in header_lower and "roll" not in header_lower:
                    headers["name"] = col_idx
                elif "roll" in header_lower:
                    headers["roll_number"] = col_idx
                elif "leetcode" in header_lower:
                    headers["leetcode_url"] = col_idx
                elif "hackerrank" in header_lower or "hacker" in header_lower:
                    headers["hackerrank_url"] = col_idx

        required = ["name", "roll_number"]
        for req_field in required:
            if req_field not in headers:
                raise HTTPException(
                    status_code=400,
                    detail=f"Column '{req_field}' not found. Expected columns: Name, Roll Number, Leetcode Profile URL, HackerRank URL",
                )

        # Delete existing data for this category uploaded by this user
        db.query(Student).filter(
            Student.category == category,
            Student.uploaded_by == current_user.id,
        ).delete()
        db.commit()

        students_added = 0
        for row in sheet.iter_rows(min_row=2, values_only=False):
            name_val = row[headers["name"] - 1].value
            if not name_val:
                continue

            roll_val = row[headers["roll_number"] - 1].value or ""
            leetcode_val = row[headers.get("leetcode_url", 0) - 1].value if headers.get("leetcode_url") else ""
            hackerrank_val = row[headers.get("hackerrank_url", 0) - 1].value if headers.get("hackerrank_url") else ""

            student = Student(
                name=str(name_val).strip(),
                roll_number=str(roll_val).strip(),
                leetcode_url=str(leetcode_val).strip() if leetcode_val else "",
                hackerrank_url=str(hackerrank_val).strip() if hackerrank_val else "",
                category=category,
                uploaded_by=current_user.id,
            )
            db.add(student)
            students_added += 1

        # Update user upload status
        if category == "1st_year":
            current_user.has_uploaded_1st_year = True
        elif category == "2nd_year":
            current_user.has_uploaded_2nd_year = True
        elif category == "3rd_year":
            current_user.has_uploaded_3rd_year = True
        elif category == "4th_year":
            current_user.has_uploaded_4th_year = True

        db.commit()
        return {"message": f"Successfully uploaded {students_added} students for {category}", "count": students_added}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


# ─── Student Data Routes ────────────────────────────────────────
@app.get("/api/students/{category}")
def get_students(
    category: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if category not in ("1st_year", "2nd_year", "3rd_year", "4th_year"):
        raise HTTPException(status_code=400, detail="Invalid category")

    students = (
        db.query(Student)
        .filter(Student.category == category, Student.uploaded_by == current_user.id)
        .order_by(Student.roll_number)
        .all()
    )

    return {
        "students": [
            {
                "id": s.id,
                "name": s.name,
                "roll_number": s.roll_number,
                "leetcode_url": s.leetcode_url,
                "hackerrank_url": s.hackerrank_url,
                "leetcode_solved": s.leetcode_solved,
                "hr_java_stars": s.hr_java_stars,
                "hr_python_stars": s.hr_python_stars,
                "hr_c_stars": s.hr_c_stars,
                "hr_sql_stars": s.hr_sql_stars,
                "last_fetched": s.last_fetched.isoformat() if s.last_fetched else None,
            }
            for s in students
        ],
        "count": len(students),
    }


@app.post("/api/students/fetch/{category}")
def fetch_student_data(
    category: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Scrape LeetCode and HackerRank data for all students in a category."""
    if category not in ("1st_year", "2nd_year", "3rd_year", "4th_year"):
        raise HTTPException(status_code=400, detail="Invalid category")

    students = (
        db.query(Student)
        .filter(Student.category == category, Student.uploaded_by == current_user.id)
        .all()
    )

    if not students:
        raise HTTPException(status_code=404, detail="No students found for this category")

    results = []
    for student in students:
        scraped = scrape_student_data(student.leetcode_url, student.hackerrank_url)
        student.leetcode_solved = scraped["leetcode_solved"]
        student.hr_java_stars = scraped["hr_java_stars"]
        student.hr_python_stars = scraped["hr_python_stars"]
        student.hr_c_stars = scraped["hr_c_stars"]
        student.hr_sql_stars = scraped["hr_sql_stars"]
        student.last_fetched = datetime.now(timezone.utc)
        results.append(
            {
                "id": student.id,
                "name": student.name,
                "leetcode_solved": scraped["leetcode_solved"],
                "hr_java_stars": scraped["hr_java_stars"],
                "hr_python_stars": scraped["hr_python_stars"],
                "hr_c_stars": scraped["hr_c_stars"],
                "hr_sql_stars": scraped["hr_sql_stars"],
            }
        )

    db.commit()
    return {"message": f"Fetched data for {len(results)} students", "results": results}


@app.post("/api/students/fetch-single/{student_id}")
def fetch_single_student(
    student_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Scrape data for a single student."""
    student = db.query(Student).filter(Student.id == student_id, Student.uploaded_by == current_user.id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    scraped = scrape_student_data(student.leetcode_url, student.hackerrank_url)
    student.leetcode_solved = scraped["leetcode_solved"]
    student.hr_java_stars = scraped["hr_java_stars"]
    student.hr_python_stars = scraped["hr_python_stars"]
    student.hr_c_stars = scraped["hr_c_stars"]
    student.hr_sql_stars = scraped["hr_sql_stars"]
    student.last_fetched = datetime.now(timezone.utc)
    db.commit()

    return {
        "id": student.id,
        "name": student.name,
        "roll_number": student.roll_number,
        "leetcode_url": student.leetcode_url,
        "hackerrank_url": student.hackerrank_url,
        "leetcode_solved": student.leetcode_solved,
        "hr_java_stars": student.hr_java_stars,
        "hr_python_stars": student.hr_python_stars,
        "hr_c_stars": student.hr_c_stars,
        "hr_sql_stars": student.hr_sql_stars,
        "last_fetched": student.last_fetched.isoformat() if student.last_fetched else None,
    }


@app.delete("/api/students/{category}")
def delete_category_data(
    category: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if category not in ("1st_year", "2nd_year", "3rd_year", "4th_year"):
        raise HTTPException(status_code=400, detail="Invalid category")

    deleted = (
        db.query(Student)
        .filter(Student.category == category, Student.uploaded_by == current_user.id)
        .delete()
    )

    if category == "1st_year":
        current_user.has_uploaded_1st_year = False
    elif category == "2nd_year":
        current_user.has_uploaded_2nd_year = False
    elif category == "3rd_year":
        current_user.has_uploaded_3rd_year = False
    elif category == "4th_year":
        current_user.has_uploaded_4th_year = False

    db.commit()
    return {"message": f"Deleted {deleted} students from {category}"}


@app.get("/api/export/{category}")
def export_category_data(
    category: str,
    min_lc: Optional[int] = Query(None),
    min_java: Optional[int] = Query(None),
    min_python: Optional[int] = Query(None),
    min_c: Optional[int] = Query(None),
    min_sql: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if category not in ("1st_year", "2nd_year", "3rd_year", "4th_year"):
        raise HTTPException(status_code=400, detail="Invalid category")

    query = db.query(Student).filter(
        Student.category == category, 
        Student.uploaded_by == current_user.id
    )

    # Apply filters if provided
    if min_lc is not None:
        query = query.filter(Student.leetcode_solved >= min_lc)
    if min_java is not None:
        query = query.filter(Student.hr_java_stars >= min_java)
    if min_python is not None:
        query = query.filter(Student.hr_python_stars >= min_python)
    if min_c is not None:
        query = query.filter(Student.hr_c_stars >= min_c)
    if min_sql is not None:
        query = query.filter(Student.hr_sql_stars >= min_sql)

    students = query.order_by(Student.roll_number).all()

    if not students:
        raise HTTPException(status_code=404, detail="No data matching filters to export")

    # Create Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{category.replace('_', ' ').title()} Data"

    # Define headers
    headers = [
        "S.No", 
        "Name", 
        "Roll Number", 
        "LeetCode Solved", 
        "Java Stars", 
        "Python Stars", 
        "C Stars", 
        "SQL Stars",
        "LeetCode URL",
        "HackerRank URL"
    ]
    ws.append(headers)

    # Add data
    for idx, s in enumerate(students, 1):
        ws.append([
            idx,
            s.name,
            s.roll_number,
            s.leetcode_solved,
            s.hr_java_stars,
            s.hr_python_stars,
            s.hr_c_stars,
            s.hr_sql_stars,
            s.leetcode_url,
            s.hackerrank_url
        ])

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{category}_coding_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/health")
def health_check():
    return {"status": "ok", "message": "Coding Retriever API is running"}

# ─── Static File Serving (Production) ───────────────────────
# Mount the frontend's 'dist' folder (created after 'npm run build')
frontend_dist_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "frontend", "dist")

if os.path.exists(frontend_dist_path):
    app.mount("/assets", StaticFiles(directory=os.path.join(frontend_dist_path, "assets")), name="assets")

    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        # If the path starts with api/, it's a 404 for the frontend catch-all
        if full_path.startswith("api/"):
            raise HTTPException(status_code=404)
        
        # Check if the requested file exists in the dist folder
        file_path = os.path.join(frontend_dist_path, full_path)
        if os.path.isfile(file_path):
            return FileResponse(file_path)
        
        # For SPA routing, return index.html for any other non-file path
        return FileResponse(os.path.join(frontend_dist_path, "index.html"))
